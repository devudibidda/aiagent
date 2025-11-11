"""
Professional Excel Report Generator
Creates formatted Excel files with compliance analysis
"""

import logging
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

EXCEL_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None
    logger.warning("openpyxl not available. Excel export disabled.")


class ComplianceReportGenerator:
    """Generate professional Excel compliance reports"""
    
    def __init__(self, output_dir: str = "./cir_output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_report(
        self,
        report_name: str,
        cir_metadata_list: List[Dict[str, Any]],
        compliance_evidence_list: List[List[Dict[str, Any]]],
        cim_requirements: List[Dict[str, Any]]
    ) -> str:
        """
        Generate comprehensive Excel report
        
        Returns:
            Path to generated Excel file
        """
        
        if not EXCEL_AVAILABLE:
            logger.warning("Excel export not available. Install: pip install openpyxl")
            return ""
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Compliance Summary"
            
            # Add title
            ws['A1'] = "VESTAS CIR COMPLIANCE ANALYSIS REPORT"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:Z1')
            
            # Add metadata
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Report: {report_name}"
            ws['A4'] = f"Total CIRs Analyzed: {len(cir_metadata_list)}"
            
            # Create summary table
            row = 6
            
            # Headers
            headers = self._get_dynamic_headers(cir_metadata_list, cim_requirements)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            row += 1
            
            # Data rows
            for i, (cir_meta, evidence) in enumerate(zip(cir_metadata_list, compliance_evidence_list)):
                col = 1
                
                # CIR metadata columns
                for header in headers:
                    if header in cir_meta:
                        cell = ws.cell(row=row, column=col)
                        cell.value = cir_meta[header]
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    elif header in ["GO/NO-GO", "Compliance Score"]:
                        # Calculate from evidence
                        pass
                    col += 1
                
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 15
            
            # Add evidence sheet
            self._add_evidence_sheet(wb, compliance_evidence_list)
            
            # Add requirements sheet
            self._add_requirements_sheet(wb, cim_requirements)
            
            # Save file
            output_file = self.output_dir / f"{report_name}_compliance_report.xlsx"
            wb.save(output_file)
            
            logger.info(f"Report generated: {output_file}")
            return str(output_file)
        
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return ""
    
    def _get_dynamic_headers(
        self,
        cir_metadata_list: List[Dict[str, Any]],
        cim_requirements: List[Dict[str, Any]]
    ) -> List[str]:
        """Generate dynamic header list based on content"""
        
        # Priority metadata fields
        priority_fields = [
            "CIR ID", "Report Type", "Turbine ID", "WTG ID",
            "Component Type", "Reason for Service",
            "Service Date", "Technician"
        ]
        
        # Get all unique fields from CIR metadata
        all_fields = set()
        for cir_meta in cir_metadata_list:
            all_fields.update(cir_meta.keys())
        
        headers = []
        
        # Add priority fields
        for field in priority_fields:
            if field in all_fields:
                headers.append(field)
        
        # Add remaining CIR fields
        for field in sorted(all_fields):
            if field not in headers:
                headers.append(field)
        
        # Add requirement columns
        if cim_requirements:
            headers.append("Compliance Score")
            headers.append("GO/NO-GO")
            headers.append("Met Requirements")
            headers.append("Partial Requirements")
            headers.append("Not Met Requirements")
        
        return headers
    
    def _add_evidence_sheet(
        self,
        wb: Workbook,
        compliance_evidence_list: List[List[Dict[str, Any]]]
    ):
        """Add detailed evidence sheet"""
        
        ws = wb.create_sheet("Evidence Details")
        
        # Headers
        headers = [
            "CIR ID", "Requirement ID", "Requirement",
            "Status", "Evidence Found", "Comments",
            "Confidence", "Visual Evidence"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data
        row = 2
        for cir_idx, cir_evidence in enumerate(compliance_evidence_list):
            for evidence in cir_evidence:
                col = 1
                for header in headers:
                    cell = ws.cell(row=row, column=col)
                    
                    if header == "CIR ID":
                        cell.value = f"CIR-{cir_idx + 1}"
                    elif header == "Status":
                        cell.value = evidence.get("Status", "")
                        # Color code status
                        if cell.value == "Met":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value == "Partial":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif cell.value == "Not Met":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        cell.value = evidence.get(header, "")
                    
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    col += 1
                
                row += 1
        
        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _add_requirements_sheet(
        self,
        wb: Workbook,
        cim_requirements: List[Dict[str, Any]]
    ):
        """Add CIM requirements reference sheet"""
        
        ws = wb.create_sheet("CIM Requirements")
        
        headers = [
            "Requirement ID", "Title", "Type", "Description",
            "Severity", "Applicable To"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add requirements
        for row, req in enumerate(cim_requirements, 2):
            ws.cell(row=row, column=1).value = req.get("id", "")
            ws.cell(row=row, column=2).value = req.get("title", "")
            ws.cell(row=row, column=3).value = req.get("type", "")
            ws.cell(row=row, column=4).value = req.get("description", "")
            ws.cell(row=row, column=5).value = req.get("severity", "")
            ws.cell(row=row, column=6).value = ", ".join(req.get("components", []))
        
        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
